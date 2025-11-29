from pythonosc.dispatcher import Dispatcher
from pythonosc.osc_server import BlockingOSCUDPServer
import threading
import tkinter as tk
from tkinter import font, messagebox
import json
from datetime import datetime
from collections import deque
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

class EEGAdaptiveReader:
    def __init__(self, root):
        self.root = root
        self.root.title("EEG-Guided Adaptive Reading Interface")
        self.root.geometry("1400x900")
        self.root.configure(bg="#0a0e27")
        
        # EEG Data Storage
        self.alpha_buffer = deque(maxlen=15)
        self.beta_buffer = deque(maxlen=15)
        self.attention_history = deque(maxlen=100)
        self.log_data = []
        self.adaptation_events = []
        
        # State Variables
        self.current_attention = 0.5
        self.prev_attention = 0.5
        self.is_focused = None
        self.font_size = 16
        self.text_color = "#e0e0e0"
        self.bg_color = "#0a0e27"
        self.focus_lower_threshold = 0.42
        self.focus_upper_threshold = 0.58
        self.session_start = datetime.now()
        self.last_adaptation = None
        
        # Memory Confidence Score
        self.memory_confidence = 0.5
        self.low_focus_start = None
        self.last_color_change = datetime.now()
        self.color_change_cooldown = 2.0  # seconds
        
        # Sample Reading Text
        self.reading_text = """You should understand that in many ways I love and respect my sister. When we were younger, Mariah was, by common agreement, the most intellectually able of my parents' four children, and the one most earnestly and touchingly devoted to the impossible work of gaining their approval. Her successes in high school and college warmed my father's heart. To warm my mother's, Mariah married once and happily, an earlier fiancé who would have been a disaster having conveniently absconded with her best friend, and she produced grandchildren with a regularity and enthusiasm that delighted my parents. Her husband is white and boring, an investment banker ten years her senior whom she met, she told the family, on a blind date, although sweet Kimmer always insists that it could only have been the personals. At Shepard Street, Mariah is greeting callers in the foyer, formal and sober in a midnight blue dress and a single strand of pearls, very much the lady of the house, as my mother might have said. From somewhere in the house wafts my father's terrible taste in classical music: Puccini with an English-language libretto.¹ The foyer is small and murky and crowded with mismatched pieces of heavy wooden furniture. It opens on the left to the

living room, on the right to the dining room, and in the back to a hallway leading to the family room and kitchen. A broad but undistinguished staircase strides upward next to the dining-room door, and along the upstairs hall is a gallery where I used to crouch in order to spy on my parents' dinner parties and poker games, and where Addison once made me hide in a successful effort to prove to me that there is no Santa Claus. Beyond the gallery is the cavernous study where my father died. To my surprise, I see two or three people up there now, leaning on the banister as though it belongs to them. In fact, there are more people in the house than I expect. The entire first floor seems filled with somber suits, a larger slice of financially comfortable African America than most white Americans probably think exists outside the sports and entertainment worlds, and I wonder how many of the guests are happier about my father's death than their faces attest. When I step through the front door, my sister offers me not a hug but a distant kiss, one cheek, other cheek, and murmurs, "I'm so glad you're here," the way she might say it to one of my father's law partners or poker buddies. Then, holding my shoulders in something still short of a hug, she looks past me down the walk, eyes tired but bright and mischievous: "Where's Kimberly?" (Mariah refuses to say Kimmer, which reeks, she once told me, of faux preppiness, although my wife attended Miss Porter's School and is thus fully qualified as a preppie.)


"On her way back from San Francisco,"I say. "She's been out there for a few days on business." Bentley, I add, much too fast, is with our neighbors: I picked him up early from his preschool yesterday and then left him again this morning to make this trip, assuming I would be too busy today to spend much time with him. Kimmer will retrieve him tonight, and they will be down tomorrow on the train. Explaining all these logistical details, knowing already that I am talking too much, I experience a yawning emptiness that I hope my face does not show, for I am missing my wife in ways I am not yet prepared to review for the family. But I need not have bothered to mask my emotions, for Mariah has plenty of her own to cope with, and makes no effort to hide her pain or her confusion. She has already forgotten asking for my wife. "I don't understand it," she says softly, shaking her head, her fingers digging into my upper arms. Actually, I am sure Mariah understands perfectly. Just last year the Judge was in the hospital to repair the imprecise results of his bypass operation of two years before, a fact my sister knows as well as I do; our father's death, if not precisely awaited, was hardly unexpected. 
"""
        
        self.setup_ui()
        self.start_osc_server()
        self.update_metrics()
    
    def setup_ui(self):
        # Header
        header_frame = tk.Frame(self.root, bg="#1a2847", height=70)
        header_frame.pack(fill=tk.X, padx=0, pady=0)
        header_frame.pack_propagate(False)
        
        title = tk.Label(header_frame, text="🧠 EEG-Guided Adaptive Reading System", 
                        font=("Helvetica", 24, "bold"), fg="#00d9ff", bg="#1a2847")
        title.pack(side=tk.LEFT, padx=25, pady=15)
        
        subtitle = tk.Label(header_frame, text="Real-time Neural Feedback Interface", 
                           font=("Helvetica", 11), fg="#7dd3fc", bg="#1a2847")
        subtitle.pack(side=tk.LEFT, padx=25)
        
        # Main Content Frame
        main_frame = tk.Frame(self.root, bg="#0a0e27")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Left Panel - Metrics
        left_panel = tk.Frame(main_frame, bg="#0f1b3f", relief=tk.FLAT)
        left_panel.pack(side=tk.LEFT, fill=tk.BOTH, padx=(0, 20), ipadx=20, ipady=20)
        left_panel.configure(width=280)
        
        metrics_label = tk.Label(left_panel, text="📊 Live Metrics", 
                                font=("Helvetica", 15, "bold"), fg="#00d9ff", bg="#0f1b3f")
        metrics_label.pack(pady=15)
        
        # Attention Level Indicator - Large Gauge
        self.attention_canvas = tk.Canvas(left_panel, width=220, height=130, 
                                         bg="#0a0e27", highlightthickness=0)
        self.attention_canvas.pack(pady=20)
        
        # Alpha/Beta Display
        alpha_frame = tk.Frame(left_panel, bg="#0f1b3f")
        alpha_frame.pack(fill=tk.X, pady=8, padx=10)
        tk.Label(alpha_frame, text="Alpha:", font=("Helvetica", 10, "bold"), 
                fg="#64b5f6", bg="#0f1b3f").pack(side=tk.LEFT)
        self.alpha_label = tk.Label(alpha_frame, text="0.0", 
                                   font=("Courier", 11, "bold"), fg="#81c784", bg="#0f1b3f")
        self.alpha_label.pack(side=tk.RIGHT)
        
        beta_frame = tk.Frame(left_panel, bg="#0f1b3f")
        beta_frame.pack(fill=tk.X, pady=8, padx=10)
        tk.Label(beta_frame, text="Beta:", font=("Helvetica", 10, "bold"), 
                fg="#ff9800", bg="#0f1b3f").pack(side=tk.LEFT)
        self.beta_label = tk.Label(beta_frame, text="0.0", 
                                  font=("Courier", 11, "bold"), fg="#ffb74d", bg="#0f1b3f")
        self.beta_label.pack(side=tk.RIGHT)
        
        focus_frame = tk.Frame(left_panel, bg="#0f1b3f")
        focus_frame.pack(fill=tk.X, pady=8, padx=10)
        tk.Label(focus_frame, text="Focus Score:", font=("Helvetica", 10, "bold"), 
                fg="#ffd700", bg="#0f1b3f").pack(side=tk.LEFT)
        self.focus_label = tk.Label(focus_frame, text="50%", 
                                   font=("Courier", 12, "bold"), fg="#ffeb3b", bg="#0f1b3f")
        self.focus_label.pack(side=tk.RIGHT)
        
        # Status Indicator
        self.status_indicator = tk.Label(left_panel, text="● Waiting for Data", 
                                        font=("Helvetica", 11, "bold"), fg="#ffaa00", bg="#0f1b3f")
        self.status_indicator.pack(pady=15)
        
        # Color preview
        color_label = tk.Label(left_panel, text="Text Color Adaptation:", 
                              font=("Helvetica", 10, "bold"), fg="#64b5f6", bg="#0f1b3f")
        color_label.pack(pady=(15, 8))
        self.color_preview = tk.Canvas(left_panel, width=200, height=30, bg="#0a0e27", 
                                       highlightthickness=2, highlightbackground="#333333")
        self.color_preview.pack()
        
        # Center Panel - Reading Interface
        center_panel = tk.Frame(main_frame, bg="#0f1b3f", relief=tk.FLAT)
        center_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, ipadx=20, ipady=20)
        
        # Memory Confidence at top of reading area
        memory_header = tk.Frame(center_panel, bg="#0f1b3f")
        memory_header.pack(fill=tk.X, pady=(0, 15))
        
        tk.Label(memory_header, text="📖 Reading Content", 
                font=("Helvetica", 13, "bold"), fg="#00d9ff", bg="#0f1b3f").pack(side=tk.LEFT)
        
        self.memory_label = tk.Label(memory_header, text="🧠 Memory Confidence: 50%", 
                                    font=("Helvetica", 13, "bold"), fg="#9c27b0", bg="#0f1b3f")
        self.memory_label.pack(side=tk.RIGHT, padx=10)
        
        reading_label = tk.Label(center_panel, text="", bg="#0f1b3f")
        reading_label.pack()
        
        # Text Display with Scrollbar
        scroll_frame = tk.Frame(center_panel, bg="#0f1b3f")
        scroll_frame.pack(fill=tk.BOTH, expand=True)
        
        scrollbar = tk.Scrollbar(scroll_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.text_display = tk.Text(scroll_frame, wrap=tk.WORD, 
                                    yscrollcommand=scrollbar.set,
                                    bg="#0a0e27", fg=self.text_color,
                                    font=("Georgia", 16),
                                    relief=tk.FLAT, bd=0, padx=20, pady=20,
                                    selectbackground="#1e3a8a")
        self.text_display.pack(fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.text_display.yview)
        
        self.text_display.insert(1.0, self.reading_text)
        self.text_display.config(state=tk.DISABLED)
        
        # Right Panel - Session Info
        right_panel = tk.Frame(main_frame, bg="#0f1b3f", relief=tk.FLAT)
        right_panel.pack(side=tk.LEFT, fill=tk.BOTH, padx=(20, 0), ipadx=20, ipady=20)
        right_panel.configure(width=280)
        
        session_label = tk.Label(right_panel, text="📈 Session Info", 
                                font=("Helvetica", 15, "bold"), fg="#00d9ff", bg="#0f1b3f")
        session_label.pack(pady=15)
        
        self.time_label = tk.Label(right_panel, text="Time: 0s", 
                                  font=("Courier", 12), fg="#e0e0e0", bg="#0f1b3f")
        self.time_label.pack(pady=8, anchor=tk.W, padx=10)
        
        self.data_points_label = tk.Label(right_panel, text="Data Points: 0", 
                                         font=("Courier", 12), fg="#e0e0e0", bg="#0f1b3f")
        self.data_points_label.pack(pady=8, anchor=tk.W, padx=10)
        
        self.focus_events_label = tk.Label(right_panel, text="Color Adaptations: 0", 
                                          font=("Courier", 12), fg="#e0e0e0", bg="#0f1b3f")
        self.focus_events_label.pack(pady=8, anchor=tk.W, padx=10)
        
        self.avg_attention_label = tk.Label(right_panel, text="Avg Attention: 0%", 
                                           font=("Courier", 12), fg="#e0e0e0", bg="#0f1b3f")
        self.avg_attention_label.pack(pady=8, anchor=tk.W, padx=10)
        
        # Adaptation Log
        tk.Label(right_panel, text="🔔 Adaptations Log", 
                font=("Helvetica", 11, "bold"), fg="#00ff88", bg="#0f1b3f").pack(pady=(20, 8))
        
        self.log_text = tk.Text(right_panel, height=18, width=30, 
                               bg="#0a0e27", fg="#81c784",
                               font=("Courier", 8), relief=tk.FLAT, bd=0)
        self.log_text.pack(fill=tk.BOTH, expand=True)
        self.log_text.config(state=tk.DISABLED)
    
    def draw_attention_gauge(self):
        self.attention_canvas.delete("all")
        
        # Background circle
        self.attention_canvas.create_oval(10, 10, 210, 110, fill="#0a0e27", outline="#1e3a8a", width=2)
        
        # Attention arc with gradient feel
        attention_pct = max(0, min(1, self.current_attention))
        
        if attention_pct < self.focus_lower_threshold:
            color = "#ef5350"  # Red
        elif attention_pct > self.focus_upper_threshold:
            color = "#66bb6a"  # Green
        else:
            color = "#ffa726"  # Orange (transition zone)
        
        angle = int(180 * attention_pct)
        self.attention_canvas.create_arc(20, 20, 200, 100, start=180, extent=angle, 
                                        fill=color, outline=color, width=4)
        
        # Center text
        text = f"{int(attention_pct * 100)}%"
        self.attention_canvas.create_text(110, 60, text=text, fill=color, 
                                         font=("Helvetica", 18, "bold"))
    
    def get_adaptive_color(self):
        """Generate color based on attention level with wider color range"""
        attention = self.current_attention
        
        if attention < 0.25:
            return "#c62828", "#ffcdd2"  # Dark red, light red
        elif attention < 0.40:
            return "#e53935", "#ef9a9a"  # Red
        elif attention < 0.50:
            return "#fb8c00", "#ffe0b2"  # Orange
        elif attention < 0.65:
            return "#fdd835", "#fff9c4"  # Yellow
        elif attention < 0.80:
            return "#7cb342", "#c5e1a5"  # Light green
        else:
            return "#2e7d32", "#81c784"  # Dark green
    
    def adapt_reading_interface(self):
        """Adjust text color based on attention level"""
        prev_color = self.text_color
        main_color, bg_tint = self.get_adaptive_color()
        self.text_color = main_color
        
        # Only trigger if color changed AND cooldown has passed
        time_since_last_change = (datetime.now() - self.last_color_change).total_seconds()
        
        if prev_color != main_color and time_since_last_change > self.color_change_cooldown:
            self.text_display.config(state=tk.NORMAL)
            self.text_display.config(fg=self.text_color)
            self.text_display.config(state=tk.DISABLED)
            
            elapsed = (datetime.now() - self.session_start).total_seconds()
            attention_pct = int(self.current_attention * 100)
            
            if self.current_attention < self.focus_lower_threshold:
                status = "🔴 Low Focus"
            elif self.current_attention > self.focus_upper_threshold:
                status = "🟢 High Focus"
            else:
                status = "🟡 Moderate Focus"
            
            self.add_log(f"{status} - {attention_pct}% Attention")
            self.last_color_change = datetime.now()
            
            self.adaptation_events.append({
                "timestamp": datetime.now().isoformat(),
                "elapsed_seconds": elapsed,
                "attention_before": self.prev_attention,
                "attention_after": self.current_attention,
                "text_color": main_color,
                "status": status
            })
    
    def add_log(self, message):
        """Add message to adaptation log"""
        self.log_text.config(state=tk.NORMAL)
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
    
    def alpha_handler(self, address, *args):
        """Handle alpha frequency data from Muse"""
        try:
            alpha_val = float(args[0])
            self.alpha_buffer.append(alpha_val)
            self.calculate_attention()
        except (IndexError, ValueError):
            pass
    
    def beta_handler(self, address, *args):
        """Handle beta frequency data from Muse"""
        try:
            beta_val = float(args[0])
            self.beta_buffer.append(beta_val)
            self.calculate_attention()
        except (IndexError, ValueError):
            pass
    
    def calculate_attention(self):
        """Calculate attention score from alpha/beta ratio"""
        if len(self.alpha_buffer) < 5 or len(self.beta_buffer) < 5:
            return
        
        alpha_avg = np.mean(list(self.alpha_buffer))
        beta_avg = np.mean(list(self.beta_buffer))
        
        # Beta-dominant indicates focus; Alpha-dominant indicates relaxation
        if beta_avg + alpha_avg > 0:
            attention_score = beta_avg / (beta_avg + alpha_avg + 0.001)
        else:
            attention_score = 0.5
        
        self.prev_attention = self.current_attention
        self.current_attention = attention_score
        self.attention_history.append(attention_score)
        
        # Evaluate low-focus duration for penalty
        if self.current_attention < 0.40:
            if self.low_focus_start is None:
                self.low_focus_start = datetime.now()
        else:
            self.low_focus_start = None
        
        # Compute average + stability
        att_list = list(self.attention_history)
        avg_att = np.mean(att_list)
        std_att = np.std(att_list)
        
        penalty = 1.0
        if self.low_focus_start:
            duration = (datetime.now() - self.low_focus_start).total_seconds()
            if duration > 3:
                penalty = 0.7
        
        self.memory_confidence = max(0, min(1, avg_att * (1 - std_att) * penalty))
        
        # Trigger adaptation
        self.adapt_reading_interface()
        
        self.log_data.append({
            "timestamp": datetime.now().isoformat(),
            "attention": attention_score,
            "alpha": alpha_avg,
            "beta": beta_avg
        })
    
    def start_osc_server(self):
        """Start OSC server in background thread"""
        disp = Dispatcher()
        disp.map("/muse/elements/alpha_absolute", self.alpha_handler)
        disp.map("/muse/elements/beta_absolute", self.beta_handler)
        
        server = BlockingOSCUDPServer(("0.0.0.0", 9001), disp)
        
        def run_server():
            try:
                server.serve_forever()
            except:
                pass
        
        thread = threading.Thread(target=run_server, daemon=True)
        thread.start()
        
        self.add_log("✓ OSC Server started on 0.0.0.0:9001")
    
    def save_session_data(self):
        """Save session data to Excel files"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Save raw EEG data
        if self.log_data:
            eeg_file = f"eeg_data_{timestamp}.xlsx"
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "EEG Data"
                
                # Headers
                headers = ["Timestamp", "Attention", "Alpha", "Beta"]
                ws.append(headers)
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                
                # Data rows
                for entry in self.log_data:
                    ws.append([entry["timestamp"], entry["attention"], entry["alpha"], entry["beta"]])
                
                # Auto-adjust column widths
                ws.column_dimensions['A'].width = 25
                ws.column_dimensions['B'].width = 12
                ws.column_dimensions['C'].width = 12
                ws.column_dimensions['D'].width = 12
                
                wb.save(eeg_file)
                print(f"✓ Saved {len(self.log_data)} EEG data points to {eeg_file}")
            except Exception as e:
                print(f"Error saving EEG data: {e}")
        else:
            print("⚠ No EEG data collected yet")
        
        # Save adaptations
        if self.adaptation_events:
            adaptations_file = f"eeg_adaptations_{timestamp}.xlsx"
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Adaptations"
                
                # Headers
                headers = ["Timestamp", "Elapsed Seconds", "Attention Before", "Attention After", "Text Color", "Status"]
                ws.append(headers)
                header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                
                # Data rows
                for entry in self.adaptation_events:
                    ws.append([entry["timestamp"], entry["elapsed_seconds"], 
                             entry["attention_before"], entry["attention_after"],
                             entry["text_color"], entry["status"]])
                
                # Auto-adjust column widths
                ws.column_dimensions['A'].width = 25
                ws.column_dimensions['B'].width = 16
                ws.column_dimensions['C'].width = 18
                ws.column_dimensions['D'].width = 18
                ws.column_dimensions['E'].width = 12
                ws.column_dimensions['F'].width = 15
                
                wb.save(adaptations_file)
                print(f"✓ Saved {len(self.adaptation_events)} adaptations to {adaptations_file}")
            except Exception as e:
                print(f"Error saving adaptations: {e}")
        else:
            print("⚠ No adaptation events recorded yet")
    
    def update_metrics(self):
        """Update UI metrics every 100ms"""
        elapsed = (datetime.now() - self.session_start).total_seconds()
        self.time_label.config(text=f"Time: {int(elapsed)}s")
        self.data_points_label.config(text=f"Data Points: {len(self.attention_history)}")
        self.focus_events_label.config(text=f"Color Adaptations: {len(self.adaptation_events)}")
        
        if len(self.attention_history) > 0:
            avg_attention = np.mean(list(self.attention_history))
            self.avg_attention_label.config(text=f"Avg Attention: {int(avg_attention * 100)}%")
        
        self.memory_label.config(text=f"🧠 Memory Confidence: {int(self.memory_confidence * 100)}%")
        
        if len(self.alpha_buffer) > 0:
            self.alpha_label.config(text=f"{np.mean(list(self.alpha_buffer)):.2f}")
            self.status_indicator.config(text="● Data Streaming", fg="#81c784")
        
        if len(self.beta_buffer) > 0:
            self.beta_label.config(text=f"{np.mean(list(self.beta_buffer)):.2f}")
        
        self.focus_label.config(text=f"{int(self.current_attention * 100)}%")
        self.draw_attention_gauge()
        
        # Update color preview
        main_color, _ = self.get_adaptive_color()
        self.color_preview.delete("all")
        self.color_preview.create_rectangle(0, 0, 200, 30, fill=main_color, outline="")
        
        self.root.after(100, self.update_metrics)
    
    def on_closing(self):
        """Handle window close event"""
        self.save_session_data()
        self.root.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    app = EEGAdaptiveReader(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()